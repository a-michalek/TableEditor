import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import os

# Read the HTML file
input_folder = r'C:\Users\Mottaka\Desktop\Current Guests'
input_file = os.path.join(input_folder, 'currentguests.xls')

tables = pd.read_html(input_file)

# Assuming you want to work with the first table
df = tables[0]

# 1. Remove the first two rows
df = df.iloc[2:].reset_index(drop=True)  # Keep all rows after the second

# Flatten MultiIndex columns if necessary
if isinstance(df.columns, pd.MultiIndex):
    df.columns = [' '.join(str(i) for i in col).strip() for col in df.columns]  # Corrected

# 2. Remove the 9th column
df = df.drop(df.columns[8], axis=1)  # 0-indexed, so 8 is the 9th column

# 3. Remove the 5th column
df = df.drop(df.columns[4], axis=1)  # 0-indexed, so 4 is the 5th column

# 4. Replace the header row with the specified column names
new_column_names = ['Arrival', 'Departure', 'Unit', 'Room', 'Guest Name', 'Occupancy', 'Duration']
df.columns = new_column_names  # Set new column names

# 5. Sort the data in ascending order by the 'Unit' column
df.sort_values(by='Unit', ascending=True, inplace=True)  # Sort in place

# 6. Determine the title and filename based on the current time
now = datetime.now()
shift_date = now - timedelta(days=1) if now.hour < 8 else now  # Set to yesterday's date if nightshift
formatted_date = shift_date.strftime('%d/%m/%Y')  # Format date as DD/MM/YYYY
title = f'Nightshift - {formatted_date}' if now.hour < 8 else f'Dayshift - {formatted_date}'

# Create a new DataFrame for the title with the same number of columns as df
title_df = pd.DataFrame([[title] + [''] * (df.shape[1] - 1)], columns=df.columns)

# Concatenate the title DataFrame and the main DataFrame
df_final = pd.concat([title_df, df], ignore_index=True)

# Define the output filename with the appropriate date
output_file_name = f'modified_file_{shift_date.strftime("%Y-%m-%d")}.xlsx'  # Date format for filename
output_file_path = os.path.join(input_folder, output_file_name)

# Save the modified DataFrame to Excel
df_final.to_excel(output_file_path, index=False, header=True)  # Save with headers

# Adjust column widths and format the title row
wb = load_workbook(output_file_path)  # Load the workbook
ws = wb.active  # Get the active worksheet

# Set the width of each column based on the maximum length of data in each column
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Add extra space for better visibility
    ws.column_dimensions[column[0].column_letter].width = adjusted_width  # Set the width

# Merge the title cell and make it bold
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(new_column_names))  # Merge cells in the title row
ws['A2'].font = Font(bold=True)  # Set the font of the merged cell to bold
ws['A2'].alignment = Alignment(horizontal='center')  # Center the text in merged cell

# Save the workbook after adjusting widths and formatting
wb.save(output_file_path)